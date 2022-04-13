from returns.result import Result, Success, Failure
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

class ExcelManager:
    def create_xlsx(
                template_file="",
                destination_file="",
                pipe_line="",
                spool="",
                drawing="",
                lst_measurements="",
                date="",
                name=""
        )->Result[bool,str]:
        try:
            wb = load_workbook(filename=template_file)
        except Exception as e:
            return Failure(f"Failed to open Excel file. {str(e)}")
        
        sheet = wb.active
        sheet["N1"] = spool
        sheet["G5"] = drawing
        sheet["N5"] = f"{drawing}/DIM1"
        sheet["C5"] = pipe_line

        for i, ms in enumerate(lst_measurements):
            sheet[f"A{17+i}"] = spool
            sheet[f"E{17+i}"] = ms[0]  #id
            sheet[f"G{17+i}"] = ms[1].upper()  #type
            sheet[f"H{17+i}"] = ms[3]  #nominal
            sheet[f"K{17+i}"] = ms[2]  #actual
            sheet[f"N{17+i}"] = f"=ABS(K{17+i}-H{17+i})"  #abs(int(ms[3])-int(ms[2]))
            sheet[f"P{17+i}"] = ms[4]  #added
            sheet[f"R{17+i}"] = "Y"  
            sheet[f"T{17+i}"] = "2" if ms[1].upper()=="A" else "1"  # "A" or "L" 
            sheet[f"U{17+i}"] = name
            sheet[f"X{17+i}"] = date

        try:
            wb.save(filename=destination_file)
        except Exception as e:
            return Failure("Failed to save xlsx file. It might exist already and be opened. " + str(e))

        return Success(True)

    
    def convert_xlsx_to_pdf(xlsx_path: str, pdf_file_name: str)-> Result[bool, str]:
        from win32com import client
        
        excel = client.Dispatch("Excel.Application")
        
        infile = os.path.abspath(xlsx_path)
        outfile = os.path.join(os.path.dirname(infile), pdf_file_name)

        sheets = excel.Workbooks.Open(infile)
        work_sheets = sheets.Worksheets[0]
        
        try:
            work_sheets.ExportAsFixedFormat(0, outfile)
        except Exception as e:
            return Failure("Failed to convert Excel to PDF. PDF might exist already and be opened. " + str(e))
        finally:
            excel.Quit()
        
        return Success(True)


    def fill_stiki_po_spulam_xlsx(self, path_stiki_xlsx: str, part_names: list, date: str)-> Result[bool, str]:
        try:
            wb = load_workbook(path_stiki_xlsx)
        except Exception as e:
            return Failure(f"Failed to open Excel file. {str(e)}")
            
        ws = wb.active
        header = ws[1]

        col_part = self.find_column_containing_str(header, 'part no')
        col_isometric = self.find_column_containing_str_excluding_str(header, 'isometric', 'draw')
        col_date = self.find_column_containing_str(header, 'date')

        for part in part_names:
            row_part = self.find_row_containing_str(ws[col_part], part.name_with_zeros())
            if row_part:
                ws[f'{col_isometric}{row_part}'] = 'OK'
                ws[f'{col_date}{row_part}'] = date

        try:
            wb.save(filename=path_stiki_xlsx)
        except Exception as e:
            return Failure("Failed to save xlsx file. It might exist already and be opened. " + str(e))
        
        return Success(True)

    
    def find_column_containing_str(self, row, s: str):
        for cell in row:
            if cell.value is not None: 
                if s.lower() in cell.value.lower():
                    return get_column_letter(cell.column)
        return None
    
    def find_row_containing_str(self, col, s: str):
        for cell in col:
            if cell.value is not None: #We need to check that the cell is not empty.
                if s.lower() in cell.value.lower(): #Check if the value of the cell contains the string
                    return cell.row
        return None
    
    def find_column_containing_str_excluding_str(self, row, s_cont: str, s_excl: str):
        for cell in row:
            if cell.value is not None:
                if s_cont.lower() in cell.value.lower() and s_excl.lower() not in cell.value.lower():
                    return get_column_letter(cell.column)
        return None

    '''
    def generate_converted_xml_thread(file=""):
        try:
            wb = load_workbook(filename=file)
        except Exception as e:
            print("Failed to open an Excel file. " + str(e))
            return
        
        sheet = wb.active
        try:
            sheet["S18"] = "L1C"
            sheet["U18"] = "L1D"
            sheet["S19"] = "R"
            sheet["T19"] = "R"
            sheet["U19"] = "R"
            sheet["V19"] = "R"

        except Exception as e:
            print("Failed to write to a file. " + str(e))
            return

        try:
            new_file = file[0:-9]+"thread.xlsx"
            wb.save(filename=new_file)
            print("Success")
        except Exception as e:
            print("Failed to save the file. It might exist already and be opened. " + str(e))
            return
    '''